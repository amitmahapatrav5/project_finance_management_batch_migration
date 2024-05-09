from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException, SheetTitleException

from pydantic import ValidationError

from typing import List
from pathlib import Path

from .constants import Constants
from .schemas import Category, SubCategory, Expense

import logging

logger = logging.getLogger(__name__)

def get_category_sheet_data() -> List[Category]:
    data = []
    try:
        wb = load_workbook(filename=Constants.FILE_PATH.value, read_only=True)
    except FileNotFoundError as e:
        logger.error(f"File not found: {Constants.FILE_PATH.value}")
        raise
    except InvalidFileException as e:
        logger.error(f"Invalid file format: {Constants.FILE_PATH.value}")
        raise

    try:
        sheet = wb[Constants.CATEGORY_SHEET_NAME.value]
    except KeyError as e:
        logger.error(f"Sheet not found: {Constants.CATEGORY_SHEET_NAME.value}")
        raise

    for row in sheet.iter_rows(values_only=True, min_row=2):
        try:
            category_data = {"category_name": row[0]}
            category = Category(**category_data)
            data.append(category)
        except Exception as e:
            logger.exception(f"There is an issue with the Category {row[0]}")
    return data

def get_sub_category_sheet_data() -> List[SubCategory]:
    data = []
    try:
        wb = load_workbook(filename=Constants.FILE_PATH.value, read_only=True)
    except FileNotFoundError as e:
        logger.error(f"File not found: {Constants.FILE_PATH.value}")
        raise
    except InvalidFileException as e:
        logger.error(f"Invalid file format: {Constants.FILE_PATH.value}")
        raise

    try:
        sheet = wb[Constants.SUB_CATEGORY_SHEET_NAME.value]
    except KeyError as e:
        logger.error(f"Sheet not found: {Constants.SUB_CATEGORY_SHEET_NAME.value}")
        raise

    for row in sheet.iter_rows(values_only=True, min_row=2):
        try:
            sub_category_data = {
                "category_name": row[0],
                "sub_category_name": row[1]
            }
            sub_category = SubCategory(**sub_category_data)
            data.append(sub_category)
        except Exception as e:
            logger.exception(f"There is an issue with the Sub Category {row[1]} of Category {row[0]}")
    return data

def get_expense_sheet_data() -> List[Expense]:
    data = []
    try:
        wb = load_workbook(filename=Constants.FILE_PATH.value, read_only=True)
    except FileNotFoundError as e:
        logger.error(f"File not found: {Constants.FILE_PATH.value}")
        raise
    except InvalidFileException as e:
        logger.error(f"Invalid file format: {Constants.FILE_PATH.value}")
        raise

    try:
        sheet = wb[Constants.EXPENSE_SHEET_NAME.value]
    except KeyError as e:
        logger.error(f"Sheet not found: {Constants.EXPENSE_SHEET_NAME.value}")
        raise

    for row in sheet.iter_rows(values_only=True, min_row=2):
        expense_data = {
            "date": row[0],
            "category_name": row[1],
            "sub_category_name": row[2],
            "product_or_service": row[3],
            "product_or_service_name": row[4],
            "quantity_or_duration": row[5],
            "price": row[6],
            "need_or_want": row[7],
            "details": row[8]
        }
        
        try:
            expense = Expense(**expense_data)
            data.append(expense)
        except ValidationError as exc:
            logger.error(f"Validation error for the expense data {expense_data}")
            logger.exception(exc)
    return data